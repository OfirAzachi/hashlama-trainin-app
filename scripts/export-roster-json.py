"""
Reads the fitness-test roster Excel export and writes scripts/roster-import.json,
which import-roster.mjs then upserts into public.roster.

Usage: python scripts/export-roster-json.py "<path to .xlsx>"
"""
import json
import re
import sys

import openpyxl

if len(sys.argv) < 2:
    print('Usage: python scripts/export-roster-json.py "<path to .xlsx>"')
    sys.exit(1)

path = sys.argv[1]
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb[wb.sheetnames[0]]

headers = [c.value for c in ws[1]]


def parse_time_to_seconds(value):
    if not isinstance(value, str):
        return None
    m = re.match(r'^(\d{1,2}):(\d{2}):(\d{2})$', value.strip())
    if not m:
        return None
    h, mi, s = m.groups()
    return int(h) * 3600 + int(mi) * 60 + int(s)


def grade(value):
    return value if value in ('V', 'X', 'חסר') else None


def as_int(value):
    if value is None or value == '':
        return None
    try:
        return round(float(value))
    except (TypeError, ValueError):
        return None


def as_str_or_none(value):
    if value is None or value == '':
        return None
    return str(value)


def km_levels(status_text):
    """Parses 'כמ 0/1/2' mentions out of the free-text status column — a
    person can carry more than one at once, e.g. 'כמ 1 - עבר, כמ 0'."""
    if not status_text:
        return []
    return sorted({int(level) for level in re.findall(r'כמ\s*([0-2])', str(status_text))})


rows = []
invalid = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    record = dict(zip(headers, row))
    personal_number = str(record.get('מ.א') or '').strip()
    if not personal_number.isdigit():
        invalid.append(record)
        continue

    rows.append({
        'personal_number': personal_number,
        'last_name': str(record.get('שם משפחה') or '').strip(),
        'first_name': str(record.get('שם פרטי') or '').strip(),
        'team': as_int(record.get('צוות')),
        'gender': str(record.get('מין (ז/נ)') or '').strip(),
        'run_start_time': as_str_or_none(record.get('זמן הזנקה')),
        'run_finish_time': as_str_or_none(record.get('זמן סיום')),
        'final_run_seconds': parse_time_to_seconds(record.get('זמן ריצה סופי')),
        'run_grade': grade(record.get('הערכה ריצה')),
        'pushup_achievement': as_int(record.get('הישג (סמיכה)')),
        'strength_grade': grade(record.get('הערכה כוח')),
        'final_score': as_int(record.get('ציון סופי')),
        'final_grade': grade(record.get('הערכה סופית')),
        'unit': str(record.get('יחידה') or '').strip(),
        'status_notes': as_str_or_none(record.get('סטטוס / נדרש להשלים')),
        'km_levels': km_levels(record.get('סטטוס / נדרש להשלים')),
    })

if invalid:
    print(f'WARNING: skipped {len(invalid)} rows with an invalid/missing מ.א: {invalid[:5]}')

with open('scripts/roster-import.json', 'w', encoding='utf-8') as f:
    json.dump(rows, f, ensure_ascii=False, indent=2)

print(f'Wrote {len(rows)} rows to scripts/roster-import.json')
